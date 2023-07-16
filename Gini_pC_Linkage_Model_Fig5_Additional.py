# -*- coding: utf-8 -*-
"""
Parse and plot inequality data

Created on Fri May 19 14:54:16 2023

@author: spauliuk
"""

### load modules
import os
import sys
import openpyxl
import pylab
import numpy as np
import matplotlib.pyplot as plt   

### data location
datapath = 'C:\\Users\\spauliuk.AD\\FILES\\ARBEIT\\\PROJECTS\\Sustainable_Consumption\\Gini_pC_Linkage\\DLS_Lorenz_Model'
datafile = 'Lorenz_Curves_DLS_Pauliuk_Workbook_SI.xlsx'

### parse data
Dfile   = openpyxl.load_workbook(os.path.join(datapath,datafile), data_only=True)
Psheet  = Dfile['Population']
Isheet  = Dfile['Income_Data']

# Define data lists and arrays
P_rlabels = []
P_value   = []
P_valuea  = []
P_valueb  = []
R_labels  = []
R_labelsa = []
R_labelsb = []
P_valuex  = []
P_valuey  = []
R_labelsx = []
R_labelsy = []
R_labelsz = []
IncData   = np.zeros((10,211))

# Get data
for m in range(4,277): # population
    P_rlabels.append(Psheet.cell(m,6).value)
    P_value.append(Psheet.cell(m,8).value * 1000) # unit: 1

for m in range(3,214): # income
    s = Isheet.cell(1,m).value
    R_labels.append(s[s.rfind('\n')+1::])
YearIndices = [21,43,65,87,109,131,153,175,197,219]    
for m in range(0,10):
    for n in range(0,211):
        IncData[m,n] = Isheet.cell(YearIndices[m],n+3).value   
                

# Income: Select the countries to plot: (a): available population data, (b) Complete income data (no nan).
DelPos = []
for m in range(0,211):
    try: 
        P_rlabels.index(R_labels[m])
        R_labelsa.append(R_labels[m])
        P_valuea.append(P_value[P_rlabels.index(R_labels[m])])
    except: 
        print(R_labels[m])
        DelPos.append(m)

IncData_a = np.delete(IncData,DelPos,1)

DelPos = []
for n in range(0,182):
    DelFlag = False
    for m in range(0,10):
        if IncData_a[m,n] == 0:
            DelFlag = True
        if  np.isnan(IncData_a[m,n]) == True:
            DelFlag = True
    if DelFlag is True:
        DelPos.append(n)
    else:
        R_labelsb.append(R_labelsa[n])
        P_valueb.append(P_valuea[n])
        
IncData_b = np.delete(IncData_a,DelPos,1)        
   

### constants and calc
G_min   = 0.25
G_max   = 0.33
mls_min = 2.51
mls_max = 3.13
mdr_min = 4.19
mdr_max = 6.22

pop     = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
G_25    = [0, 0.061259607, 0.125310341, 0.192655625, 0.263978077, 0.340246045, 0.422920038, 0.514406625, 0.619269212, 0.748811357, 1]
G_33    = [0, 0.051692394, 0.106322816, 0.164461085, 0.226889437, 0.294733411, 0.369719341, 0.45475095, 0.555484109, 0.686497794, 1]
G_30    = [0, 0.055153299, 0.113216334, 0.174739132, 0.240473413, 0.311495274, 0.389445255, 0.477062355, 0.579630173, 0.710573388, 1]

G_25_t  = 1 - 2 * np.trapz(G_25,x = pop)
G_33_t  = 1 - 2 * np.trapz(G_33,x = pop)

d_LC_25 = G_25[1]/0.1 # LC slope for lowest decile: roughly the dls: 0.6 of the average.
d_LC_33 = G_33[1]/0.1 # LC slope for lowest decile: roughly the dls: 0.5 of the average.
dLC_30min = G_30[1]/0.1 # representative DLS slope
dLC_30max = (1-G_30[-2])/0.1

# Calculate Lorenz Curves, GNI, pC_GDP, and Gini Coefficients for income:
LC_Total = IncData_b.cumsum(axis=0)
LC_Norm  = np.divide(LC_Total,np.einsum('a,b->ba',LC_Total[-1,:],np.ones(10)))
LC_Norm  = np.insert(LC_Norm, 0, 0, axis=0) # Add 0 as starting point of the Lorenz curve

GNI      = np.einsum('ab,b->b',IncData_b,np.asarray(P_valueb)/10) / 1e9  # ppp GNI, 2021 constant billion EUR
pC_GNI   = np.divide(GNI,P_valueb) * 1e9

G_emp = np.zeros((169))
for m in range(0,169):
    G_emp[m]  = 1 - 2 * np.trapz(LC_Norm[:,m],x = pop)

# ratio of income for highest vs. lowest decile    
maxminr = IncData_b[-1,:] / IncData_b[0,:]

# DLS gap and exessive consumption as fraction of GNI.
LC_Norm_Slope = np.diff(LC_Norm, axis=0) * 10
LC_Norm_Slope_Comp = LC_Norm_Slope > dLC_30min
LC_Norm_Slope_Pos  = 10 - LC_Norm_Slope_Comp.sum(axis=0) # convert comparison to position
LC_Norm_Slope_Val  = [LC_Norm[LC_Norm_Slope_Pos[i],i] for i in range(0,169)] # convert position to value for data
LC_Model_Slope_Val = [G_30[LC_Norm_Slope_Pos[i]] for i in range(0,169)] # convert position to value for model curve
DLS_Gap = [LC_Model_Slope_Val[i] - LC_Norm_Slope_Val[i] for i in range(0,169)] # Difference betwee both is the DLS gap as fraction of GNI
ExcessIncome = 0.1 * (LC_Norm_Slope[-1,:] - dLC_30max)

### plot data
fig, ax1 = plt.subplots(1, 1, sharex=False, gridspec_kw={'hspace': 0.3, 'wspace': 0.35},figsize=(18,15))

# scatter plot of maxmin_ratio empirical vs. model
ax1.scatter(pC_GNI, maxminr)
ax1.fill_between([0,1e5], [mdr_min,mdr_min], [mdr_max,mdr_max], linestyle = '-', facecolor = '0.75', linewidth = 0.5)
ax1.set_title('Scatter plot of empirical (dots) and model-based \n (range) top 10% to bottom 10% income ratios.', fontsize = 15)
ax1.set_ylabel('top 10% to bottom 10% income ratio', fontsize = 15)
ax1.set_xlabel('per capita GNI, in 2021 EUR', fontsize = 15)
ax1.legend(['Countries','DLS Model range'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax1.set_xlim([0, 75000])
ax1.set_ylim([0, 850])


plt.xticks(fontsize=18)
plt.yticks(fontsize=18)
plt.show()

fig.savefig(os.path.join(datapath,'Fig5_additional.png'), dpi=200, bbox_inches='tight')




# End


