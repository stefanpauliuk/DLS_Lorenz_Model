# -*- coding: utf-8 -*-
"""
Parse and plot inequality data

Created on Fri May 19 14:54:16 2023

@author: spauliuk
"""

### load modules
import os
import sys
import openpyxl
import pylab
import numpy as np
import matplotlib.pyplot as plt   

### data location
datapath = 'C:\\Users\\spauliuk.AD\\FILES\\ARBEIT\\Projects\\Sustainable_Consumption\\Gini_pC_Linkage\\DLS_Lorenz_Model'
datafile = 'Lorenz_Curves_DLS_Pauliuk_Workbook_SI.xlsx'

### parse data
Dfile   = openpyxl.load_workbook(os.path.join(datapath,datafile), data_only=True)
Dsheet  = Dfile['Lorenz_Model_Properties_Fig3']

G   = np.zeros((51,1)) # unit: 1
dpa = np.zeros((51,1)) 
dpb = np.zeros((51,1)) 
dpc = np.zeros((51,1)) 

# Get data
for m in range(0,51):
    G[m,0]    = Dsheet.cell(m+3,2).value
    dpa[m,0]  = Dsheet.cell(m+3,4).value
    dpb[m,0]  = Dsheet.cell(m+3,5).value
    dpc[m,0]  = Dsheet.cell(m+3,8).value
        
### constants and calc
None

### plot data
fig, ((ax1, ax2, ax3)) = plt.subplots(1, 3, sharex=True, gridspec_kw={'hspace': 0.3, 'wspace': 0.35},figsize=(15,5))

# Line plot of ratios.

ax1.set_prop_cycle('color', np.array([[0.45, 0.61, 0.78, 1]]))
a1 = ax1.plot(G, dpa, linestyle = '-', linewidth = 2)
ax1.set_title('(a) pcs to dls_r ratio.', fontsize = 15)
ax1.set_ylabel('Ratio pcs to dls_r', fontsize = 18)
ax1.set_xlabel('Gini coefficient G', fontsize = 18)
ax1.grid()
plt.xticks(fontsize=18)
plt.yticks(fontsize=18)
ax1.legend([a1[0]],['ratio'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax1.set_xlim([0, 0.8])
ax1.set_ylim([0, 10])

ax2.set_prop_cycle('color', np.array([[0.45, 0.61, 0.78, 1]]))
a2 = ax2.plot(G, dpb, linestyle = '-', linewidth = 2)
ax2.set_title('(b) mls to dls_r ratio.', fontsize = 15)
ax2.set_ylabel('Ratio mls to dls_r', fontsize = 18)
ax2.set_xlabel('Gini coefficient G', fontsize = 18)
ax2.grid()
plt.xticks(fontsize=18)
plt.yticks(fontsize=18)
ax2.legend([a2[0]],['ratio'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax2.set_xlim([0, 0.8])
ax2.set_ylim([0, 80])

ax3.set_prop_cycle('color', np.array([[0.45, 0.61, 0.78, 1]]))
a3 = ax3.plot(G, dpc, linestyle = '-', linewidth = 2)
ax3.set_title('(c) median to dls_r ratio.', fontsize = 15)
ax3.set_ylabel('Ratio median to dls_r', fontsize = 18)
ax3.set_xlabel('Gini coefficient G', fontsize = 18)
ax3.grid()
plt.xticks(fontsize=18)
plt.yticks(fontsize=18)
ax3.legend([a2[0]],['ratio'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax3.set_xlim([0, 1])
ax3.set_ylim([0.5, 1.2])

plt.show()
fig.savefig(os.path.join(datapath,'Fig3.png'), dpi=400, bbox_inches='tight')



# End


