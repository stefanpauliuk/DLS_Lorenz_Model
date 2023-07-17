# -*- coding: utf-8 -*-
"""
Parse and plot inequality data

Created on Fri May 19 14:54:16 2023

@author: spauliuk
"""

### load modules
import os
import sys
import openpyxl
import pylab
import numpy as np
import matplotlib.pyplot as plt   

### data location
datapath = 'C:\\Users\\spauliuk.AD\\FILES\\ARBEIT\\\PROJECTS\\Sustainable_Consumption\\Gini_pC_Linkage\\DLS_Lorenz_Model'
datafile = 'Lorenz_Curves_DLS_Pauliuk_Workbook_SI.xlsx'

### parse data
Dfile   = openpyxl.load_workbook(os.path.join(datapath,datafile), data_only=True)
Psheet  = Dfile['Population']
Isheet  = Dfile['Income_Data']
Wsheet  = Dfile['Wealth_Data']

# Define data lists and arrays
P_rlabels = []
P_value   = []
P_valuea  = []
P_valueb  = []
R_labels  = []
R_labelsa = []
R_labelsb = []
P_valuex  = []
P_valuey  = []
R_labelsx = []
R_labelsy = []
R_labelsz = []
IncData   = np.zeros((10,211))
WealthData= np.zeros((10,207))

# Get data
for m in range(4,277): # population
    P_rlabels.append(Psheet.cell(m,6).value)
    P_value.append(Psheet.cell(m,8).value * 1000) # unit: 1

for m in range(3,214): # income
    s = Isheet.cell(1,m).value
    R_labels.append(s[s.rfind('\n')+1::])
YearIndices = [21,43,65,87,109,131,153,175,197,219]    
for m in range(0,10):
    for n in range(0,211):
        IncData[m,n] = Isheet.cell(YearIndices[m],n+3).value   
        
for m in range(3,210): # wealth
    s = Wsheet.cell(1,m).value
    R_labelsx.append(s[s.rfind('\n')+1::])
YearIndices = [11,23,35,47,59,71,83,95,107,119]    
for m in range(0,10):
    for n in range(0,207):
        WealthData[m,n] = Wsheet.cell(YearIndices[m],n+3).value   
        

# Income: Select the countries to plot: (a): available population data, (b) Complete income data (no nan).
DelPos = []
for m in range(0,211):
    try: 
        P_rlabels.index(R_labels[m])
        R_labelsa.append(R_labels[m])
        P_valuea.append(P_value[P_rlabels.index(R_labels[m])])
    except: 
        print(R_labels[m])
        DelPos.append(m)

IncData_a = np.delete(IncData,DelPos,1)

DelPos = []
for n in range(0,182):
    DelFlag = False
    for m in range(0,10):
        if IncData_a[m,n] == 0:
            DelFlag = True
        if  np.isnan(IncData_a[m,n]) == True:
            DelFlag = True
    if DelFlag is True:
        DelPos.append(n)
    else:
        R_labelsb.append(R_labelsa[n])
        P_valueb.append(P_valuea[n])
        
IncData_b = np.delete(IncData_a,DelPos,1)        

# Wealth: Select the countries to plot: (a): available population data, (b) Complete income data (no nan).
DelPos = []
for m in range(0,207):
    try: 
        P_rlabels.index(R_labelsx[m])
        R_labelsy.append(R_labelsx[m])
        P_valuex.append(P_value[P_rlabels.index(R_labelsx[m])])
    except: 
        print(R_labelsx[m])
        DelPos.append(m)

WealthData_a = np.delete(WealthData,DelPos,1)

DelPos = []
for n in range(0,181):
    DelFlag = False
    for m in range(0,10):
        if  np.isnan(WealthData_a[m,n]) == True:
            DelFlag = True
        if  R_labelsy[n] == 'Venezuela': # implausibly high values for Venezuela, remove from sample
            DelFlag = True            
    if DelFlag is True:
        DelPos.append(n)
        print(R_labelsy[n])
    else:
        R_labelsz.append(R_labelsy[n])
        P_valuey.append(P_valuex[n])
        
WealthData_b = np.delete(WealthData_a,DelPos,1)     

### constants and calc
G_min   = 0.25
G_max   = 0.33
mls_min = 2.51
mls_max = 3.13
mdr_min = 4.19
mdr_max = 6.22

pop     = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
G_25    = [0, 0.061259607, 0.125310341, 0.192655625, 0.263978077, 0.340246045, 0.422920038, 0.514406625, 0.619269212, 0.748811357, 1]
G_33    = [0, 0.051692394, 0.106322816, 0.164461085, 0.226889437, 0.294733411, 0.369719341, 0.45475095, 0.555484109, 0.686497794, 1]
G_30    = [0, 0.055153299, 0.113216334, 0.174739132, 0.240473413, 0.311495274, 0.389445255, 0.477062355, 0.579630173, 0.710573388, 1]

G_25_t  = 1 - 2 * np.trapz(G_25,x = pop)
G_33_t  = 1 - 2 * np.trapz(G_33,x = pop)

d_LC_25 = G_25[1]/0.1 # LC slope for lowest decile: roughly the dls: 0.6 of the average.
d_LC_33 = G_33[1]/0.1 # LC slope for lowest decile: roughly the dls: 0.5 of the average.
dLC_30min = G_30[1]/0.1 # representative DLS slope
dLC_30max = (1-G_30[-2])/0.1

# Calculate Lorenz Curves, GNI, pC_GDP, and Gini Coefficients for income:
LC_Total = IncData_b.cumsum(axis=0)
LC_Norm  = np.divide(LC_Total,np.einsum('a,b->ba',LC_Total[-1,:],np.ones(10)))
LC_Norm  = np.insert(LC_Norm, 0, 0, axis=0) # Add 0 as starting point of the Lorenz curve

GNI      = np.einsum('ab,b->b',IncData_b,np.asarray(P_valueb)/10) / 1e9  # ppp GNI, 2021 constant billion EUR
pC_GNI   = np.divide(GNI,P_valueb) * 1e9

G_emp = np.zeros((169))
for m in range(0,169):
    G_emp[m]  = 1 - 2 * np.trapz(LC_Norm[:,m],x = pop)

# ratio of income for highest vs. lowest decile    
maxminr = IncData_b[-1,:] / IncData_b[0,:]

# DLS gap and exessive consumption as fraction of GNI.
LC_Norm_Slope = np.diff(LC_Norm, axis=0) * 10
LC_Norm_Slope_Comp = LC_Norm_Slope > dLC_30min
LC_Norm_Slope_Pos  = 10 - LC_Norm_Slope_Comp.sum(axis=0) # convert comparison to position
LC_Norm_Slope_Val  = [LC_Norm[LC_Norm_Slope_Pos[i],i] for i in range(0,169)] # convert position to value for data
LC_Model_Slope_Val = [G_30[LC_Norm_Slope_Pos[i]] for i in range(0,169)] # convert position to value for model curve
DLS_Gap = [LC_Model_Slope_Val[i] - LC_Norm_Slope_Val[i] for i in range(0,169)] # Difference betwee both is the DLS gap as fraction of GNI
ExcessIncome = 0.1 * (LC_Norm_Slope[-1,:] - dLC_30max)

# Calculate Lorenz Curves, GNI, pC_GDP, and Gini Coefficients for wealth:
LC_Total_W = WealthData_b.cumsum(axis=0)
LC_Norm_W  = np.divide(LC_Total_W,np.einsum('a,b->ba',LC_Total_W[-1,:],np.ones(10)))
LC_Norm_W  = np.insert(LC_Norm_W, 0, 0, axis=0) # Add 0 as starting point of the Lorenz curve

NPW        = np.einsum('ab,b->b',WealthData_b,np.asarray(P_valuey)/10) / 1e9  # ppp net personal wealth, 2021 constant billion EUR
pC_NPW     = np.divide(NPW,P_valuey) * 1e9

G_emp_W = np.zeros((180))
for m in range(0,180):
    G_emp_W[m]  = 1 - 2 * np.trapz(LC_Norm_W[:,m],x = pop)

### plot data
ccolors = ['#BF8F00']
fig, ((ax1, ax2, ax3), (ax4, ax5, ax6)) = plt.subplots(2, 3, sharex=False, gridspec_kw={'hspace': 0.3, 'wspace': 0.35},figsize=(28,15))

# Line plot of Lorenz curves
ax1.set_prop_cycle('color', np.array([[0.72, 0.8, 0.89, 1]]))
a1 = ax1.plot(pop,LC_Norm)
a2 = ax1.plot(pop,G_25,color = 'g', linewidth = 1.5)
ax = ax1.plot([0,1], [0,1], linestyle = '--', color ='k', linewidth = 1)
a3 = ax1.plot(pop,G_33,color = 'g', linewidth = 1.5)
ax1.set_title('(a) Empirical (blue) vs model-based \n (green) Lorenz curves for income.', fontsize = 15)
ax1.set_ylabel('Cumulative fraction of total income, %.', fontsize = 15)
ax1.set_xlabel('Cumulative fraction of total population, %.', fontsize = 15)
ax1.legend([a1[0],a3[0],ax[0]],['Countries','DLS Models for G = 0.25 and 0.33','Equal distribution'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax1.set_xlim([0, 1])
ax1.set_ylim([0, 1])
ax1.grid()

# scatter plot of G_emp vs G_model
ax2.scatter(pC_GNI, G_emp)
ax2.fill_between([0,1e5], [G_min,G_min], [G_max,G_max], linestyle = '-', facecolor = '0.75', linewidth = 0.5)
ax2.set_title('(b) Scatter plot of empirical (dots) and \n model-based (range) Gini coefficients for income.', fontsize = 15)
ax2.set_ylabel('Gini coefficient', fontsize = 15)
ax2.set_xlabel('per capita GNI, in 2021 EUR', fontsize = 15)
ax2.legend(['Countries','DLS Model range'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax2.set_xlim([0, 75000])
ax2.set_ylim([0.2, 0.8])
ax2.grid()

# plot DLS gap and excessive consumption as fraction of GNI
ax3.scatter(pC_GNI, np.asarray(DLS_Gap))
ax3.scatter(pC_GNI, ExcessIncome, color = ccolors[0])
ax3.set_title('(c) DLS gap and exessive consumption \n as fraction of GNI with G = 0.30 as reference.', fontsize = 15)
ax3.set_ylabel('Fraction of GNI', fontsize = 15)
ax3.set_xlabel('per capita GNI, in 2021 EUR', fontsize = 15)
ax3.legend(['DLS gap for G = 0.3','Excessive income'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax3.set_xlim([0, 75000])
ax3.set_ylim([0, 0.4])
ax3.grid()

# Line plot of Lorenz curves for wealth
ax4.set_prop_cycle('color', np.array([[0.72, 0.8, 0.89, 1]]))
a1 = ax4.plot(pop,LC_Norm_W)
ax = ax4.plot([0,1], [0,1], linestyle = '--', color ='k', linewidth = 1)
ax4.set_title('(d) Empirical Lorenz curves for wealth.', fontsize = 15)
ax4.set_ylabel('Cumulative fraction of total income, %.', fontsize = 15)
ax4.set_xlabel('Cumulative fraction of total population, %.', fontsize = 15)
ax4.legend([a1[0],ax[0]],['Countries','Equal distribution'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax4.set_xlim([0, 1])
ax4.set_ylim([0, 1])
ax4.grid()

# scatter plot of G_emp_W
ax5.scatter(pC_NPW/1000, G_emp_W)
ax5.set_title('(e) Scatter plot of empirical Gini coefficients for wealth.', fontsize = 15)
ax5.set_ylabel('Gini coefficient', fontsize = 15)
ax5.set_xlabel('per capita net pers. wealth, in 2021 kEUR', fontsize = 15)
ax5.legend(['Countries'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax5.set_xlim([-20, 400])
ax5.set_ylim([0.6, 0.92])
ax5.grid()

# DLS gap to exc. cons. ratio
ax = ax6.scatter(pC_GNI, np.asarray(DLS_Gap)/ExcessIncome)
ax6.plot([0,75000], [1,1], linestyle = '--', color ='k', linewidth = 1)
ax6.set_title('(f) DLS gap to exessive consumption ratio, \n with G = 0.30 as reference.', fontsize = 15)
ax6.set_ylabel('Ratio', fontsize = 15)
ax6.set_xlabel('per capita GNI, in 2021 EUR', fontsize = 15)
ax6.legend([ax],['DLS gap to Excessive income ratio'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper right')# ,bbox_to_anchor=(1.91, 1)) 
ax6.set_xlim([0, 75000])
ax6.set_ylim([0.5, 3])
ax6.grid()

plt.xticks(fontsize=16)
plt.yticks(fontsize=16)
plt.show()

fig.savefig(os.path.join(datapath,'Fig5.png'), dpi=200, bbox_inches='tight')

# Export central results to allow for country identification

rbook = openpyxl.Workbook() # Export other model results, calibration values, flags, etc.
wsx = rbook.active
wsx.title = 'Results'

# Income data
wsx.cell(row=1, column=2).value = 'Regions'
wsx.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)

wsx.cell(row=1, column=3).value = 'Lorenz curve (normalized) for income, Fig. 5(a)'
wsx.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)

for i in range(0,169):
    wsx.cell(row=i+2, column=2).value = R_labelsb[i]
for i in range(0,11):
    wsx.cell(row=1, column=4+i).value = i*0.1
    for j in range(0,169):
        wsx.cell(row=j+2, column=4+i).value = LC_Norm[i,j]
        
wsx.cell(row=1, column=16).value = 'per capita GNI, 2021 EUR, Fig. 5(b)'
wsx.cell(row=1, column=16).font = openpyxl.styles.Font(bold=True)
wsx.cell(row=1, column=17).value = 'Gini coefficient for income, Fig. 5(b)'
wsx.cell(row=1, column=17).font = openpyxl.styles.Font(bold=True)  
wsx.cell(row=1, column=18).value = 'ratio of income for highest vs. lowest decile, Fig. 5 (additonal result)'
wsx.cell(row=1, column=18).font = openpyxl.styles.Font(bold=True)  
wsx.cell(row=1, column=19).value = 'DLS gap as fraction of GNI, Fig. 5(c)'
wsx.cell(row=1, column=19).font = openpyxl.styles.Font(bold=True) 
wsx.cell(row=1, column=20).value = 'Exessive consumption as fraction of GNI, Fig. 5(c)'
wsx.cell(row=1, column=20).font = openpyxl.styles.Font(bold=True) 
wsx.cell(row=1, column=21).value = 'DLS gap to exessive consumption ratio, Fig. 5(f)'
wsx.cell(row=1, column=21).font = openpyxl.styles.Font(bold=True) 

for i in range(0,169):
    wsx.cell(row=i+2, column=16).value = pC_GNI[i]
    wsx.cell(row=i+2, column=17).value = G_emp[i]
    wsx.cell(row=i+2, column=18).value = maxminr[i]
    wsx.cell(row=i+2, column=19).value = np.asarray(DLS_Gap)[i]
    wsx.cell(row=i+2, column=20).value = ExcessIncome[i]
    wsx.cell(row=i+2, column=21).value = np.asarray(DLS_Gap)[i]/ExcessIncome[i]
    
# Wealth data    
wsx.cell(row=1, column=25).value = 'Regions'
wsx.cell(row=1, column=25).font = openpyxl.styles.Font(bold=True)

wsx.cell(row=1, column=26).value = 'Empirical Lorenz curves for wealth, Fig. 5(d)'
wsx.cell(row=1, column=26).font = openpyxl.styles.Font(bold=True)

for i in range(0,180):
    wsx.cell(row=i+2, column=25).value = R_labelsz[i]
for i in range(0,11):
    wsx.cell(row=1, column=27+i).value = i*0.1
    for j in range(0,180):
        wsx.cell(row=j+2, column=27+i).value = LC_Norm_W[i,j]    

wsx.cell(row=1, column=40).value = 'Per capita net pers. wealth, in 2021 kEUR, Fig. 5(e)'
wsx.cell(row=1, column=40).font = openpyxl.styles.Font(bold=True) 
wsx.cell(row=1, column=41).value = 'Gini coefficient for wealth, Fig. 5(e)'
wsx.cell(row=1, column=41).font = openpyxl.styles.Font(bold=True)     
    
for i in range(0,180):
    wsx.cell(row=i+2, column=40).value = pC_NPW[i]
    wsx.cell(row=i+2, column=41).value = G_emp_W[i]

    
rbook.save(os.path.join(datapath,'AddResults.xlsx'))

# End


