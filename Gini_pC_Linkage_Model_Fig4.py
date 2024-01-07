# -*- coding: utf-8 -*-
"""
Parse and plot inequality data

Created on Fri May 19 14:54:16 2023

@author: spauliuk
"""

### load modules
import os
import openpyxl
import numpy as np
import matplotlib.pyplot as plt   

### data location
datapath  = 'C:\\Users\\spauliuk.AD\\FILES\\ARBEIT\\Projects\\Sustainable_Consumption\\Gini_pC_Linkage\\DLS_Lorenz_Model'
datafile  = 'Lorenz_Curves_DLS_Pauliuk_Workbook_SI.xlsx'

### parse data
Dfile   = openpyxl.load_workbook(os.path.join(datapath,datafile), data_only=True)
Dsheet  = Dfile['pav_reb_nrb_Global_Fig4']

P   = np.zeros((20)) # unit: million
S   = np.zeros((20,3)) # per capita stock for reb (m² per person), nrb (m² per person), and pav (vehicles per person)
E   = np.zeros((20,3)) # Total use phase final energy consumption for reb (TJ/yr), nrb (TJ/yr), and pav (TJ/yr)

# Get data
for m in range(0,20):
    P[m]      = Dsheet.cell(m+4,5).value
    S[m,0]    = Dsheet.cell(m+4,8).value  # reb
    S[m,1]    = Dsheet.cell(m+4,12).value # nrb
    S[m,2]    = Dsheet.cell(m+4,15).value # pav
    E[m,0]    = Dsheet.cell(m+4,19).value / Dsheet.cell(m+4,5).value # reb, convert to per capita energy consumption
    E[m,1]    = Dsheet.cell(m+4,21).value / Dsheet.cell(m+4,5).value # nrb, convert to per capita energy consumption
    E[m,2]    = Dsheet.cell(m+4,23).value / Dsheet.cell(m+4,5).value # pav, convert to per capita energy consumption

### constants and calc
P_share      = P/sum(P)

# LC test
S_test       = np.array([0,0,20,0,0,0,0,0,0,10,0,0,0,0,0,0,0,0,80,0])
sortx_tes_S  = np.argsort(S_test)
LC_P_tes_S   = np.insert(P_share[sortx_tes_S].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_tes_S     = np.insert(np.multiply(P[sortx_tes_S],S_test[sortx_tes_S]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_tes_S_l    = 1 - 2*np.multiply(LC_tes_S[0:-1],P_share[sortx_tes_S]).sum()/LC_tes_S[-1] # Gini coefficient for reb stock, global, by country, lower sum of the integral.
G_tes_S_u    = 1 - 2*np.multiply(LC_tes_S[1::],P_share[sortx_tes_S]).sum()/LC_tes_S[-1] # Gini coefficient for reb stock, global, by country, upper sum of the integral.
G_tex        = 1 - 2 * np.trapz(LC_tes_S,dx = P_share[sortx_tes_S]) / LC_tes_S[-1] # Gini coefficient with trapez sum, average of the two above.

# LC reb stock
sortx_reb_S  = np.argsort(S[:,0])
LC_P_reb_S   = np.insert(P_share[sortx_reb_S].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_reb_S     = np.insert(np.multiply(P[sortx_reb_S],S[sortx_reb_S,0]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_reb_S_l    = 1 - 2*np.multiply(LC_reb_S[0:-1],P_share[sortx_reb_S]).sum()/LC_reb_S[-1] # Gini coefficient for reb stock, global, by country.
G_reb_S_u    = 1 - 2*np.multiply(LC_reb_S[1::],P_share[sortx_reb_S]).sum()/LC_reb_S[-1] # Gini coefficient for reb stock, global, by country.
G_reb_S      = 1 - 2 * np.trapz(LC_reb_S,dx = P_share[sortx_reb_S]) / LC_reb_S[-1] # Gini coefficient with trapez sum, average of the two above.

# LC nrb stock
sortx_nrb_S  = np.argsort(S[:,1])
LC_P_nrb_S   = np.insert(P_share[sortx_nrb_S].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing nonresidential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_nrb_S     = np.insert(np.multiply(P[sortx_nrb_S],S[sortx_nrb_S,1]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing nonresidential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_nrb_S_l    = 1 - 2*np.multiply(LC_nrb_S[0:-1],P_share[sortx_nrb_S]).sum()/LC_nrb_S[-1] # Gini coefficient for nrb stock, global, by country.
G_nrb_S_u    = 1 - 2*np.multiply(LC_nrb_S[1::],P_share[sortx_nrb_S]).sum()/LC_nrb_S[-1] # Gini coefficient for nrb stock, global, by country.
G_nrb_S      = 1 - 2 * np.trapz(LC_nrb_S,dx = P_share[sortx_nrb_S]) / LC_nrb_S[-1] # Gini coefficient with trapez sum, average of the two above.

# LC pav stock
sortx_pav_S  = np.argsort(S[:,2])
LC_P_pav_S   = np.insert(P_share[sortx_pav_S].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing car ownwership per capita. Add leading 0 (start of the Lorenz Curve).
LC_pav_S     = np.insert(np.multiply(P[sortx_pav_S],S[sortx_pav_S,2]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing car ownership per capita. Add leading 0 (start of the Lorenz Curve).
G_pav_S      = 1 - 2 * np.trapz(LC_pav_S,dx = P_share[sortx_pav_S]) / LC_pav_S[-1] # Gini coefficient with trapez sum, average of the two above.

# LC reb energy
sortx_reb_E  = np.argsort(E[:,0])
LC_P_reb_E   = np.insert(P_share[sortx_reb_E].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_reb_E     = np.insert(np.multiply(P[sortx_reb_E],E[sortx_reb_E,0]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_reb_E      = 1 - 2 * np.trapz(LC_reb_E,dx = P_share[sortx_reb_E]) / LC_reb_E[-1] # Gini coefficient with trapez sum, average of the two above.

# LC nrb energy
sortx_nrb_E  = np.argsort(E[:,1])
LC_P_nrb_E   = np.insert(P_share[sortx_nrb_E].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_nrb_E     = np.insert(np.multiply(P[sortx_nrb_E],E[sortx_nrb_E,1]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_nrb_E      = 1 - 2 * np.trapz(LC_nrb_E,dx = P_share[sortx_nrb_E]) / LC_nrb_E[-1] # Gini coefficient with trapez sum, average of the two above.

# LC pav energy
sortx_pav_E  = np.argsort(E[:,2])
LC_P_pav_E   = np.insert(P_share[sortx_pav_E].cumsum(), 0, 0) # determine cumulative sum of population shares, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
LC_pav_E     = np.insert(np.multiply(P[sortx_pav_E],E[sortx_pav_E,2]).cumsum(), 0, 0) # determine cumulative sum of stock, sorted by increasing residential floorspace per capita. Add leading 0 (start of the Lorenz Curve).
G_pav_E      = 1 - 2 * np.trapz(LC_pav_E,dx = P_share[sortx_pav_E]) / LC_pav_E[-1] # Gini coefficient with trapez sum, average of the two above.


# generate model curve for G_reb_global with 15 m²/cap
P_supprel = np.linspace(0,1,101)
u_Gl      = (1 + G_reb_S) / (1 - G_reb_S)
pCS_15_Gl = u_Gl * 15
S_Gl_15   = sum(P) * pCS_15_Gl
LC_Gl_15  = S_Gl_15 * ( 1 - np.power(1 - P_supprel, 1/u_Gl))

### plot data
ccolors = ['#BF8F00']
fig, ((ax1), (ax3), (ax4)) = plt.subplots(3, 1, sharex=False, gridspec_kw={'hspace': 0.3, 'wspace': 0.35},figsize=(10,25))

# Line plot of Lorenz curves, residential buildings global (reb)
ax1.set_prop_cycle('color', np.array([[0.72, 0.8, 0.89, 1]]))
a1 = ax1.plot(LC_P_reb_S, LC_reb_S/1000, linestyle = '-', color ='b', linewidth = 2)
a2 = ax1.plot([0,1], [0,LC_reb_S[-1]/1000], linestyle = '--', color ='k', linewidth = 1)
a3 = ax1.plot(P_supprel,LC_Gl_15/1000,linestyle = '--', color = ccolors[0], linewidth = 1.5)
ax1.set_title('(a) Residential buildings (global, 2015).', fontsize = 15)
ax1.set_ylabel('Floorspace (billion m²)', fontsize = 15)
ax1.set_xlabel('Cumulative population share', fontsize = 15)
ax1.grid()

ax1a = ax1.twinx()
a1a  = ax1a.plot(LC_P_reb_E, LC_reb_E, linestyle = '-', color ='g', linewidth = 2)
ax1a.set_ylabel('Final Energy use phase (TJ/yr)', fontsize = 15)
ax1a.set_xlim([0, 1])
ax1a.set_ylim([0,LC_reb_E[-1]])

ax1.legend([a1[0],a1a[0],a3[0]],['Floorspace (stock), G = ' + '{0:.2f}'.format(G_reb_S),'Final energy (flow), G = ' + '{0:.2f}'.format(G_reb_E),'Model curve dls = 15 m²/cap'], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax1.set_xlim([0, 1])
ax1.set_ylim([0,LC_reb_S[-1]/1000])

# Line plot of Lorenz curves, non-residential buildings global (nrb)
ax3.set_prop_cycle('color', np.array([[0.72, 0.8, 0.89, 1]]))
b1 = ax3.plot(LC_P_nrb_S, LC_nrb_S/1000, linestyle = '-', color ='b', linewidth = 2)
b2 = ax3.plot([0,1], [0,LC_nrb_S[-1]/1000], linestyle = '--', color ='k', linewidth = 1)
ax3.set_title('(b) Non-residential buildings (global, 2015).', fontsize = 15)
ax3.set_ylabel('Floorspace (billion m²)', fontsize = 15)
ax3.set_xlabel('Cumulative population share', fontsize = 15)
ax3.grid()

ax3a = ax3.twinx()
a3a  = ax3a.plot(LC_P_nrb_E, LC_nrb_E, linestyle = '-', color ='g', linewidth = 2)
ax3a.set_ylabel('Final Energy use phase (TJ/yr)', fontsize = 15)
ax3a.set_xlim([0, 1])
ax3a.set_ylim([0,LC_nrb_E[-1]])

ax3.legend([b1[0],a3a[0]],['Floorspace (stock), G = ' + '{0:.2f}'.format(G_nrb_S),'Final energy (flow), G = ' + '{0:.2f}'.format(G_nrb_E)], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax3.set_xlim([0, 1])
ax3.set_ylim([0,LC_nrb_S[-1]/1000])

# Line plot of Lorenz curves, passenver vehicles global (pav)
ax4.set_prop_cycle('color', np.array([[0.72, 0.8, 0.89, 1]]))

c1 = ax4.plot(LC_P_pav_S, LC_pav_S, linestyle = '-', color ='b', linewidth = 2)
c2 = ax4.plot([0,1], [0,LC_pav_S[-1]], linestyle = '--', color ='k', linewidth = 1)
ax4.set_title('(c) Passenger vehicles (global, 2015).', fontsize = 15)
ax4.set_ylabel('Passenger vehicles (million items)', fontsize = 15)
ax4.set_xlabel('Cumulative population share', fontsize = 15)
ax4.grid()

ax4a = ax4.twinx()
a4a  = ax4a.plot(LC_P_pav_E, LC_pav_E, linestyle = '-', color ='g', linewidth = 2)
ax4a.set_ylabel('Final Energy use phase (TJ/yr)', fontsize = 15)
ax4a.set_xlim([0, 1])
ax4a.set_ylim([0,LC_pav_E[-1]])

ax4.legend([c1[0],a4a[0]],['Passenger vehicles (stock), G = ' + '{0:.2f}'.format(G_pav_S),'Final energy (flow), G = ' + '{0:.2f}'.format(G_pav_E)], shadow = False, prop={'size':14}, ncol=1, loc = 'upper left')# ,bbox_to_anchor=(1.91, 1)) 
ax4.set_xlim([0, 1])
ax4.set_ylim([0,LC_pav_S[-1]])

plt.xticks(fontsize=15)
plt.yticks(fontsize=15)


plt.show()
fig.savefig(os.path.join(datapath,'Fig4.png'), dpi=400, bbox_inches='tight')



# End



